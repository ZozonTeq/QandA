from multiprocessing.dummy import Array
from typing import List
import openpyxl



wb = openpyxl.load_workbook("q.xlsx")
ws = wb.worksheets[0]  # アクティブなワークシートを選択
score = 0
miss = 0
failed = ""
failed_i =""
print(f'sheet name: {ws.title}')  # sheet name: Sheet
try:
    for i in range(1,64):
        if ws.cell(i,1).value is not None:
            print("========="+str(i)+"=========")
            print("[Q]"+str(ws.cell(i,2).value))
            inp =input()
            print(str(ws.cell(i,1).value))
            inp = input()
            msg = ""
            if inp == "o":
                msg ="[O]"
                score += 1
            else :
                msg ="[X]"
                miss += 1
                failed+=" , "+str(ws.cell(i,1).value)
                failed_i += " , " + str(i)
            print(msg+ "[A]"+str(ws.cell(i,1).value))
            print("score : "+str(score)+"  miss : "+str(miss))
    print(failed)
    print(failed_i)
    inp = input()


except KeyboardInterrupt:
    print(failed)
    print(failed_i)
